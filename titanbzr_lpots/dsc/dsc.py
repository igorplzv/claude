import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from scipy.signal import savgol_filter
import openpyxl
import os  # Добавлен импорт модуля os

# Настройка стиля для публикационного качества
plt.style.use('default')
plt.rcParams.update({
    'font.family': 'Arial',  # Используем универсальный шрифт для публикаций
    'font.size': 12,
    'axes.labelsize': 12,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'legend.fontsize': 10,
    'axes.grid': True,
    'grid.alpha': 0.3,
    'axes.axisbelow': True,
    'axes.spines.top': False,  # Убираем верхнюю и правую границу для минимализма
    'axes.spines.right': False,
    'lines.linewidth': 1.5,
})

def process_dsc_data(workbook):
    """Обработка данных ДСК из Excel-файла NETZSCH"""
    sheet = workbook.active
    
    # Поиск строки с заголовком данных (содержит "##Temp./°C")
    header_row = None
    for row in sheet.iter_rows():
        if row[0].value == "##Temp./°C":
            header_row = row[0].row
            break
    
    if header_row is None:
        raise ValueError("Не удалось найти строку с заголовком данных")
    
    # Извлечение данных, начиная со строки после заголовков
    temperatures = []
    dsc_signals = []
    
    for row in sheet.iter_rows(min_row=header_row + 1):
        temp = row[0].value  # Температура в первом столбце
        dsc = row[2].value   # Сигнал ДСК в третьем столбце
        
        # Пропускаем нечисловые значения
        if isinstance(temp, (int, float)) and isinstance(dsc, (int, float)):
            temperatures.append(temp)
            dsc_signals.append(dsc)
    
    return np.array(temperatures), np.array(dsc_signals)

def create_dsc_plot():
    try:
        # Чтение файлов с использованием openpyxl
        heating_wb = openpyxl.load_workbook('TiTaNbZr_Heating1.xlsx')
        cooling_wb = openpyxl.load_workbook('TiTaNbZr_Cooling1.xlsx')
        
        # Обработка данных
        heating_temp, heating_signal = process_dsc_data(heating_wb)
        cooling_temp, cooling_signal = process_dsc_data(cooling_wb)
        
        # Сглаживание данных с помощью фильтра Савитцкого-Голая
        heating_smooth = savgol_filter(heating_signal, 51, 3)
        cooling_smooth = savgol_filter(cooling_signal, 51, 3)
        
        # Создание фигуры
        fig, ax = plt.subplots(figsize=(8, 6), dpi=300)
        
        # Построение кривых
        heating_line = ax.plot(heating_temp, heating_smooth, 
                             label='Heating, 10 K/min', color='#2166AC', linewidth=1.5)
        cooling_line = ax.plot(cooling_temp, cooling_smooth, 
                             label='Cooling, 10 K/min', color='#B2182B', linewidth=1.5)
        
        # Добавление стрелок для направления нагрева/охлаждения
        ax.annotate('', xy=(750, 0.15), xytext=(700, 0.15),
                    arrowprops=dict(arrowstyle='->', color='#2166AC', lw=1.0))
        ax.annotate('', xy=(700, 0.05), xytext=(750, 0.05),
                    arrowprops=dict(arrowstyle='->', color='#B2182B', lw=1.0))
        
        # Отметки фазовых превращений для нагрева
        ax.axvline(x=815, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        ax.text(815, 0.22, 'T$_{s}^{h}$ = 815°C', rotation=90, va='bottom', ha='center', fontsize=8)
        ax.axvline(x=862, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        ax.text(862, 0.22, 'T$_{f}^{h}$ = 862°C', rotation=90, va='bottom', ha='center', fontsize=8)
        
        # Отметки фазовых превращений для охлаждения
        ax.axvline(x=804, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        ax.text(804, -0.08, 'T$_{s}^{c}$ = 804°C', rotation=90, va='top', ha='center', fontsize=8)
        ax.axvline(x=743, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        ax.text(743, -0.08, 'T$_{f}^{c}$ = 743°C', rotation=90, va='top', ha='center', fontsize=8)
        
        # Подписи фазовых превращений для сплава Ti-10Ta-2Nb-2Zr
        ax.annotate('α+β → β', xy=(840, 0.05), xytext=(840, 0.15),
                    arrowprops=dict(arrowstyle='->', color='black', lw=0.8),
                    ha='center', va='bottom', fontsize=8)
        ax.annotate('β → α+β', xy=(770, -0.05), xytext=(770, -0.15),
                    arrowprops=dict(arrowstyle='->', color='black', lw=0.8),
                    ha='center', va='top', fontsize=8)
        
        # Настройка осей и пределов
        ax.set_xlabel('Temperature (°C)')
        ax.set_ylabel('Heat Flow (mW/mg)')
        ax.set_xlim(600, 1000)
        ax.set_ylim(-0.1, 0.25)
        
        # Добавление легенды
        ax.legend(loc='upper right', frameon=True, edgecolor='black', facecolor='white', framealpha=0.8)
        
        # Добавление текста 'exo ↓' для указания экзотермического направления
        ax.text(0.02, 0.98, 'exo ↓', transform=ax.transAxes, 
                verticalalignment='top', fontsize=10, color='black')
        
        # Улучшение компоновки
        plt.tight_layout()
        
        # Сохранение графика в разные форматы
        output_dir = 'output'  # Папка для сохранения (создаётся, если не существует)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_path_png = os.path.join(output_dir, 'DSC_Ti-10Ta-2Nb-2Zr_Curves.png')
        output_path_pdf = os.path.join(output_dir, 'DSC_Ti-10Ta-2Nb-2Zr_Curves.pdf')
        
        plt.savefig(output_path_png, dpi=300, bbox_inches='tight')
        plt.savefig(output_path_pdf, bbox_inches='tight')
        
        print("График успешно создан и сохранён!")
        
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_dsc_plot()