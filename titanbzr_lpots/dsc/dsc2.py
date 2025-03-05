import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from scipy.signal import savgol_filter
import openpyxl
from matplotlib.patches import Rectangle

# Set style for publication-quality figures
plt.style.use('default')
plt.rcParams.update({
    'font.family': 'Arial',
    'font.size': 18,
    'axes.labelsize': 18,
    'xtick.labelsize': 16,
    'ytick.labelsize': 16,
    'legend.fontsize': 16,
    'axes.grid': True,
    'grid.alpha': 0.3,
    'axes.axisbelow': True,
    'axes.spines.top': True,
    'axes.spines.right': True,
    'figure.dpi': 300
})

def process_dsc_data(workbook):
    """Process NETZSCH DSC data from Excel workbook"""
    sheet = workbook.active
    
    # Find the data header row (contains "##Temp./°C")
    header_row = None
    for row in sheet.iter_rows():
        if row[0].value == "##Temp./°C":
            header_row = row[0].row
            break
    
    if header_row is None:
        raise ValueError("Could not find data header row")
    
    # Extract data starting from the row after headers
    temperatures = []
    dsc_signals = []
    
    for row in sheet.iter_rows(min_row=header_row + 1):
        temp = row[0].value  # Temperature in first column
        dsc = row[2].value   # DSC signal in third column
        
        # Skip any non-numeric values
        if isinstance(temp, (int, float)) and isinstance(dsc, (int, float)):
            temperatures.append(temp)
            dsc_signals.append(dsc)
    
    return np.array(temperatures), np.array(dsc_signals)

def create_dsc_plot():
    try:
        # Read the files using openpyxl
        heating_wb = openpyxl.load_workbook('TiTaNbZr_Heating1.xlsx')
        cooling_wb = openpyxl.load_workbook('TiTaNbZr_Cooling1.xlsx')
        
        # Process data
        heating_temp, heating_signal = process_dsc_data(heating_wb)
        cooling_temp, cooling_signal = process_dsc_data(cooling_wb)
        
        # Create figure
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Apply Savitzky-Golay filter for smoothing
        heating_smooth = savgol_filter(heating_signal, 51, 3)
        cooling_smooth = savgol_filter(cooling_signal, 51, 3)
        
        # Define colors
        heating_color = '#B2182B'  # Red
        cooling_color = '#2166AC'  # Blue
        
        # Plot the data
        heating_line = ax.plot(heating_temp, heating_smooth, 
                             label='Heating', color=heating_color, linewidth=2)
        cooling_line = ax.plot(cooling_temp, cooling_smooth, 
                             label='Cooling', color=cooling_color, linewidth=2)
        
        # Add phase transformation regions
        # Heating transformation region (815-862°C)
        ax.add_patch(Rectangle((815, -0.2), 47, 0.5,
                             facecolor=heating_color, alpha=0.1))
        # Cooling transformation region (743-804°C)
        ax.add_patch(Rectangle((743, -0.2), 61, 0.5,
                             facecolor=cooling_color, alpha=0.1))
        
        # Mark phase transformations with vertical lines and temperature labels
        props = dict(boxstyle='round', facecolor='white', alpha=0.9, edgecolor='none', pad=0.3)
        
        # Temperature labels
        # Cooling (bottom)
        ax.text(743, -0.18, "T$_{f}$ = 743°C", ha='center', va='top', size=16, color=cooling_color)
        ax.text(804, -0.18, "T$_{s}$ = 804°C", ha='center', va='top', size=16, color=cooling_color)
        # Heating (top)
        ax.text(815, 0.28, "T$_{s}$ = 815°C", ha='center', va='bottom', size=16, color=heating_color)
        ax.text(862, 0.28, "T$_{f}$ = 862°C", ha='center', va='bottom', size=16, color=heating_color)
        
        # Peak temperatures with white background
        ax.text(764, -0.173, "T$_{peak}$ = 764°C", ha='center', va='bottom', size=14, 
                color=cooling_color, bbox=dict(facecolor='white', alpha=0.9, edgecolor='none', pad=0.5))
        ax.text(842, 0.12, "T$_{peak}$ = 842°C", ha='center', va='bottom', size=14,
                color=heating_color, bbox=dict(facecolor='white', alpha=0.9, edgecolor='none', pad=0.5))
        
        # Vertical lines
        ax.axvline(x=743, color='gray', linestyle='--', alpha=0.5)
        ax.axvline(x=804, color='gray', linestyle='--', alpha=0.5)
        ax.axvline(x=815, color='gray', linestyle='--', alpha=0.5)
        ax.axvline(x=862, color='gray', linestyle='--', alpha=0.5)
        
        # Phase transformation labels with background
        ax.annotate('β → α+β', xy=(775, -0.05), xytext=(775, -0.11),
                    arrowprops=dict(arrowstyle='->', color=cooling_color),
                    ha='center', va='top', bbox=props, size=16, color=cooling_color)
        
        ax.annotate('α+β → β', xy=(840, 0.15), xytext=(840, 0.21),
                    arrowprops=dict(arrowstyle='->', color=heating_color),
                    ha='center', va='bottom', bbox=props, size=16, color=heating_color)
        
        # Customize the plot
        ax.set_xlabel('Temperature (°C)')
        ax.set_ylabel('Heat Flow (mW/mg)')
        ax.set_xlim(600, 1000)
        ax.set_ylim(-0.2, 0.3)  # Further expanded y-axis range
        
        # Add legend with better position and appearance
        ax.legend(loc='upper right', frameon=True, framealpha=1, edgecolor='none',
                 bbox_to_anchor=(0.98, 0.98), fontsize=16)
        
        # Add text 'exo ↓' to indicate exothermic direction
        ax.text(0.02, 0.98, 'exo ↓', transform=ax.transAxes, 
                verticalalignment='top', fontsize=16, bbox=props)
        
        # Adjust grid appearance
        ax.grid(True, linestyle='--', alpha=0.3)
        
        # Tight layout to prevent label clipping
        plt.tight_layout()
        
        # Save the figure
        plt.savefig('DSC_TiTaNbZr.png', dpi=300, bbox_inches='tight')
        plt.savefig('DSC_TiTaNbZr.pdf', bbox_inches='tight')
        
        print("Plot created successfully!")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_dsc_plot()